import logging
from openpyxl import Workbook

from models.document import Document

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exports accounting data into Excel with multiple sheets.
    """

    def export(self, document: Document, path: str) -> None:
        logger.info("Exporting document to Excel")

        wb = Workbook()

        # Sheet 1: Items
        ws_items = wb.active
        ws_items.title = "Items"
        ws_items.append(
            ["Description", "Quantity", "Unit Price", "Line Total", "Confidence"]
        )

        for item in document.table_items:
            ws_items.append([
                item.description,
                item.quantity,
                item.unit_price,
                item.line_total,
                item.confidence,
            ])

        # Sheet 2: VAT Breakdown
        ws_vat = wb.create_sheet("VAT")
        ws_vat.append(["Rate", "Amount"])

        if document.validation_report:
            vat_data = document.validation_report.get("vat", {})
            for rate, amount in vat_data.items():
                ws_vat.append([rate, amount])

        # Sheet 3: Validation Report
        ws_val = wb.create_sheet("Validation")
        ws_val.append(["Check", "Result"])

        if document.validation_report:
            for key, value in document.validation_report.items():
                ws_val.append([key, str(value)])

        wb.save(path)
        logger.info("Excel export saved to %s", path)
